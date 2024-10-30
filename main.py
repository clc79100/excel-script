from openpyxl import load_workbook
import os
import xlsxwriter
from xlsx_utlilites import Xlsx_Utilities
from continuous_var import Continuous_Variable
from discrete_var import Discrete_varaible
from qualitative_var import Qualitative_Variable
    
def divide_by_columns(end_table, ws, columns):
    only_column = Xlsx_Utilities.only_column_cell
    only_row = Xlsx_Utilities.only_row_cell
    to_cell = Xlsx_Utilities.column_to_only_column_cell

    end_col = only_column(end_table)
    end_row = only_row(end_table)

    #obtiene titulos de columnas
    for col in ws['A1':f'{end_col}1']:
        for cell in col:
            columns.append({'title':cell.value})
    i=0
    #obtiene los datos de cada columna en lista 
    for col in columns:
        col['data'] = [row[0].value for row in ws[f'{to_cell(i)}2':f'{to_cell(i)}{end_row}']]
        col['end_row'] = int(end_row)
        i+=1

def assign_var_to_column(columns):
    print('Ingresa tipo de Variable')
    print('Continua: c')
    print('Discreta: d')
    print('Cualitativa: q')
    print('Tamaño de clase a conveniencia: d-#')
    #itera en cada columna, para asignar tipo de clase
    for col in columns:
        var_type = input(f'- {col['title']}: ')
        if('d-' in var_type):
            col['convenience_size_classes'] = int(var_type.replace('d-',''))
            var_type = var_type[0]
        col['var_type'] = var_type.lower()

def assign_var_fun(wb, columns, c_list, d_list, q_list):
    for col in columns:
        match(col['var_type']):
            case 'c':
                c_list.append(col)
            case 'd':
                d_list.append(col)
            case 'q':
               q_list.append(col)
            case _:
                print(f'{col['title']}: None')

def loop_on_varaible(wb, c_list, d_list, q_list):
    ws_continuous_var = wb.add_worksheet('V Continua')
    ws_discrete_var = wb.add_worksheet('V Discreta')
    ws_qualitative_var = wb.add_worksheet('V Cualitativa')

    i=0
    for col in c_list:
        Continuous_Variable(wb, ws_continuous_var, col, starting_from=i)
        i+=17

    i=0
    for col in d_list:
        Discrete_varaible(wb, ws_discrete_var, col, starting_from=i)
        i+=17
    
    i=0
    for col in q_list:
        Qualitative_Variable(wb, ws_qualitative_var, col, starting_from=i)
        i+=17
        

if os.path.exists('salida.xlsx'):
    os.remove('salida.xlsx')
else:
    print('Archivo no encontrado')



columns = []
c_list, d_list, q_list = [],[],[]

external_wb =  load_workbook('IA_original.xlsx')
i=1
for sheet in external_wb:
    print(f'{i}. {sheet.title}')
    i+=1

ws_index=int(input('Ingrese num de hoja: '))

table_end = input('Ingrese fin de tabla: ').upper()

external_ws = external_wb.worksheets[ws_index-1]


divide_by_columns(table_end, external_ws, columns)

assign_var_to_column(columns)

new_workbook = xlsxwriter.Workbook('salida.xlsx')

assign_var_fun(new_workbook, columns, c_list, d_list, q_list)

loop_on_varaible(new_workbook, c_list, d_list, q_list)

new_workbook.close()
